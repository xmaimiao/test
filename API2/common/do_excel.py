from openpyxl import load_workbook
from API2.common import contants

class Cases:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.expect = None
        self.method = None
        self.sql = None

class DoExcel:
    def __init__(self,filename,sheetname):
        self.filename = filename
        self.sheetname = sheetname
        try:
            self.wb = load_workbook(self.filename)
            self.sheet = self.wb[self.sheetname]
        except FileNotFoundError as e:
            print("找不到文件！")
            raise e
        # self.cases = Cases()

    def read_excel(self):
        cases = []
        for row in range(2,self.sheet.max_row+1):
            case = Cases()
            try:
                case.case_id = self.sheet.cell(row,1).value
                case.title = self.sheet.cell(row,2).value
                case.method = self.sheet.cell(row,3).value
                case.url = self.sheet.cell(row,4).value
                case.data = self.sheet.cell(row,5).value
                case.expect = self.sheet.cell(row,6).value
                case.sql = self.sheet.cell(row,9).value
            except Exception as e:
                print("请检查取值是否正确写入列表cases！")
                raise  e
            cases.append(case)
        return cases

    def write_excel(self,row,actual,result):
        self.sheet.cell(row, 7).value = actual
        self.sheet.cell(row, 8).value = result
        self.wb.save(self.filename)
        self.wb.close()

if __name__ == '__main__':
    filename = contants.cases_dir
    data = DoExcel(filename,'sendMCode')
    r_excel = data.read_excel()
    print(type(r_excel))
    print(r_excel[0].data)
    print(eval("__import__('os').system('whoami')"))
